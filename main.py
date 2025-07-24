import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, numbers

def load_file():
    file_path = filedialog.askopenfilename()  # Możesz wybrać dowolny plik dla testu
    if file_path:
        process_file(file_path)

def process_file(file_path):
    try:
        # Wczytaj dane z Excela, ustawiając odpowiedni wiersz jako nagłówek
        df = pd.read_excel(file_path, header=1)  # Ustawienie header=1 na stałe

        # Wyświetl nagłówki kolumn, aby zweryfikować ich poprawność
        print("Columns in the file:", df.columns.tolist())

        # Wyświetl pierwsze kilka wierszy danych, aby zobaczyć ich strukturę
        print("Preview of the data:")
        print(df.head())

        # Przykład: Konwersja kolumny "Date" na format datetime
        if 'Date' not in df.columns:
            messagebox.showerror("Error", "Column 'Date' not found in the file.")
            return

        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        if df['Date'].isnull().any():
            messagebox.showerror("Error", "Some dates could not be converted.")
            print("Rows with conversion errors:")
            print(df[df['Date'].isnull()])
            return

        # Grupowanie po kampanii i dacie, zsumowanie godzin
        def calculate_hours(group):
            total_hours = 0
            take_times = group[group['Action'] == 'take']
            for index, take_action in take_times.iterrows():
                corresponding_actions = group[(group['Xdeliverable'] == take_action['Xdeliverable']) & 
                                              (group['Action'].isin(['accept', 'reject'])) &
                                              (group['Date'] > take_action['Date'])]
                if not corresponding_actions.empty:
                    accept_reject_time = corresponding_actions['Date'].iloc[0]
                    time_diff = accept_reject_time - take_action['Date']
                    # Oblicz czas w godzinach i zawsze zaokrąglij w górę do najbliższej 0.25
                    hours = np.ceil(time_diff.total_seconds() / 3600 / 0.25) * 0.25
                    hours = max(hours, 0.25)  # Minimum 0.25 godziny
                    print(f"Calculated hours between {take_action['Date']} and {accept_reject_time}: {hours}")
                    total_hours += hours
            # Ograniczanie godzin do 8 max
            return min(total_hours, 8)

        results = df.groupby(['Campaign name', df['Date'].dt.date]).apply(calculate_hours).reset_index(name='Hours')

        # Sortowanie wyników po dacie
        results = results.sort_values(by='Date')

        # Wybór miejsca zapisu pliku wynikowego
        output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if not output_file:
            return  # Jeśli użytkownik anuluje wybór

        results.to_excel(output_file, index=False)
        
        # Kolorowanie wierszy naprzemiennie
        wb = load_workbook(output_file)
        ws = wb.active

        # Definiowanie kolorów
        fill_grey = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        fill_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Czerwony

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
            current_date = row[1].value
            fill = fill_grey if current_date.weekday() % 2 == 0 else PatternFill()

            # Kolorowanie wierszy na czerwono, jeśli godziny są >= 8
            if row[2].value >= 8:
                fill = fill_red

            for cell in row:
                cell.fill = fill

        # Ustawienie formatu liczbowego z dwoma miejscami po przecinku dla kolumny "Hours"
        for cell in ws['C']:
            cell.number_format = numbers.FORMAT_NUMBER_00

        wb.save(output_file)
        print(f"Plik przetworzony i zapisany jako '{output_file}'.")
    except Exception as e:
        print(f"An error occurred: {e}")

app = tk.Tk()
app.title("CRD Processor")

# Ustawienie większego rozmiaru okna, np. 400x300 pikseli
app.geometry("400x300")

load_button = tk.Button(app, text="Load Excel File", command=load_file)
load_button.pack()

app.mainloop()